"""Excel parser supporting .xlsx (openpyxl) and .xls (xlrd)."""

import os
from pathlib import Path
from typing import Generator, Optional

from .types import ColumnInfo, ParsedDataset


# ---------------------------------------------------------------------------
# .xlsx via openpyxl
# ---------------------------------------------------------------------------

def _xlsx_row_generator(path: str, sheet_name: Optional[str], header_row: int) -> Generator:
    """Yield data rows as lists of strings from an .xlsx file."""
    import openpyxl
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[sheet_name] if sheet_name else wb.active
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == header_row:
            continue
        yield [_xlsx_cell_to_str(v) for v in row]
    wb.close()


def _xlsx_cell_to_str(value) -> str:
    if value is None:
        return ""
    if isinstance(value, bool):
        return str(value)
    if isinstance(value, float):
        return str(int(value)) if value == int(value) else str(value)
    if isinstance(value, int):
        return str(value)
    return str(value)


def _parse_xlsx(path: str, sheet: Optional[str], header_row: int) -> ParsedDataset:
    import openpyxl
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[sheet] if sheet else wb.active
    sheet_name = ws.title

    header: list = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == header_row:
            header = [str(v).strip() if v is not None else f"col_{j}" for j, v in enumerate(row)]
            break
    wb.close()

    columns = [ColumnInfo(name=name, position=i) for i, name in enumerate(header)]
    file_size = os.path.getsize(path)

    return ParsedDataset(
        columns=columns,
        row_iterator=_xlsx_row_generator(path, sheet_name, header_row),
        metadata={
            "encoding": "utf-8",
            "delimiter": None,
            "header_row": header_row,
            "sheet": sheet_name,
            "estimated_rows": 0,
            "file_size": file_size,
        },
    )


# ---------------------------------------------------------------------------
# .xls via xlrd
# ---------------------------------------------------------------------------

def _xls_cell_to_str(sheet, row_idx: int, col_idx: int, datemode: int) -> str:
    """Convert an xlrd cell to a string value."""
    import xlrd
    ctype = sheet.cell_type(row_idx, col_idx)
    value = sheet.cell_value(row_idx, col_idx)

    if ctype in (xlrd.XL_CELL_EMPTY, xlrd.XL_CELL_BLANK):
        return ""
    if ctype == xlrd.XL_CELL_ERROR:
        return ""
    if ctype == xlrd.XL_CELL_BOOLEAN:
        return str(bool(value))
    if ctype == xlrd.XL_CELL_DATE:
        try:
            dt = xlrd.xldate_as_datetime(value, datemode)
            return dt.strftime("%Y-%m-%d %H:%M:%S") if dt.hour or dt.minute or dt.second else dt.strftime("%Y-%m-%d")
        except Exception:
            return str(value)
    if ctype == xlrd.XL_CELL_NUMBER:
        if value == int(value):
            return str(int(value))
        return str(value)
    return str(value)


def _xls_row_generator(path: str, sheet_name: Optional[str], header_row: int) -> Generator:
    """Yield data rows as lists of strings from an .xls file."""
    import xlrd
    wb = xlrd.open_workbook(path)
    sh = wb.sheet_by_name(sheet_name) if sheet_name else wb.sheet_by_index(0)
    datemode = wb.datemode
    for i in range(sh.nrows):
        if i == header_row:
            continue
        yield [_xls_cell_to_str(sh, i, j, datemode) for j in range(sh.ncols)]


def _parse_xls(path: str, sheet: Optional[str], header_row: int) -> ParsedDataset:
    import xlrd
    wb = xlrd.open_workbook(path)
    sh = wb.sheet_by_name(sheet) if sheet else wb.sheet_by_index(0)
    sheet_name = sh.name

    header = [str(sh.cell_value(header_row, j)).strip() or f"col_{j}" for j in range(sh.ncols)]
    columns = [ColumnInfo(name=name, position=i) for i, name in enumerate(header)]
    file_size = os.path.getsize(path)

    return ParsedDataset(
        columns=columns,
        row_iterator=_xls_row_generator(path, sheet_name, header_row),
        metadata={
            "encoding": "utf-8",
            "delimiter": None,
            "header_row": header_row,
            "sheet": sheet_name,
            "estimated_rows": max(0, sh.nrows - 1),
            "file_size": file_size,
        },
    )


# ---------------------------------------------------------------------------
# Public entry point
# ---------------------------------------------------------------------------

def parse_excel(path: str, sheet: Optional[str] = None, header_row: int = 0) -> ParsedDataset:
    """Parse an Excel file (.xlsx or .xls) and return a ParsedDataset."""
    suffix = Path(path).suffix.lower()
    if suffix == ".xlsx":
        try:
            import openpyxl  # noqa: F401
        except ImportError:
            raise ImportError("pip install 'jdatamunch-mcp[excel]'")
        return _parse_xlsx(path, sheet, header_row)
    elif suffix == ".xls":
        try:
            import xlrd  # noqa: F401
        except ImportError:
            raise ImportError("pip install 'jdatamunch-mcp[excel]'")
        return _parse_xls(path, sheet, header_row)
    else:
        raise ValueError(f"excel_parser only handles .xlsx/.xls, got: {suffix!r}")
